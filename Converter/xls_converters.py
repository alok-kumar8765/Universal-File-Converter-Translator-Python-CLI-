# xls_converters.py
import os, sys, csv, json
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from docx import Document
import threading # Import threading for schedule_delete
from PIL import Image, ImageDraw, ImageFont
import textwrap
import math
import zipfile
import io
import unicodedata
import shutil # Import shutil for rmtree

PAGE_W, PAGE_H = A4

# ---------- Script to Font mapping (for image conversion) ----------
FALLBACK_FONTS = {
    "LATIN": "arial.ttf", # Assuming Arial is available or will be handled
    "DEVANAGARI": "NotoSansDevanagari-Medium.ttf", # Assuming Noto is available or will be handled
    "CJK": "arialuni.ttf", # Arial Unicode MS
    "ARABIC": "arial.ttf",
    "GREEK": "arial.ttf",
    "OTHER": "arialuni.ttf",
    "DEFAULT": "arial.ttf"
}

def detect_script_simple(text: str) -> str:
    """
    Heuristic script detection based on Unicode codepoints.
    Returns one of: DEVANAGARI, CJK, ARABIC, GREEK, HANGUL, LATIN, OTHER
    """
    if not text:
        return "LATIN"
    for ch in text:
        cp = ord(ch)
        if 0x0900 <= cp <= 0x097F:  # Devanagari (Hindi, etc)
            return "DEVANAGARI"
        if (0x4E00 <= cp <= 0x9FFF) or (0x3400 <= cp <= 0x4DBF) or (0x3040 <= cp <= 0x30FF):
            return "CJK"
        if 0xAC00 <= cp <= 0xD7AF:
            return "HANGUL"
        if 0x0600 <= cp <= 0x06FF or 0x0750 <= cp <= 0x077F:
            return "ARABIC"
        if 0x0370 <= cp <= 0x03FF:
            return "GREEK"
    if any(ord(ch) > 127 for ch in text):
        return "OTHER"
    return "LATIN"

def get_font_path(script, default_size=12):
    font_name = FALLBACK_FONTS.get(script, FALLBACK_FONTS["DEFAULT"])
    # In a Colab environment, commonly available fonts are in specific paths
    # You might need to adjust these paths or download fonts if needed.
    # Example paths (may vary):
    colab_font_paths = [
        f"/usr/share/fonts/truetype/liberation/{font_name}", # Liberation fonts often present
        f"/usr/share/fonts/truetype/msttcorefonts/{font_name}", # If ttf-mscorefonts-installer is run
        f"/usr/share/fonts/truetype/noto/{font_name}", # Noto fonts
        f"/content/{font_name}", # If user uploaded
        font_name # As a last resort, try just the name
    ]

    for font_path in colab_font_paths:
        if os.path.exists(font_path):
            try:
                return ImageFont.truetype(font_path, default_size)
            except Exception as e:
                print(f"⚠️ Could not load font {font_path}: {e}")

    # If specific font not found, try a generic fallback font that supports Unicode
    try:
        # Try a widely available Unicode font like Arial Unicode MS or Noto Sans
        fallback_unicode_fonts = ["arialuni.ttf", "NotoSans-Regular.ttf"]
        for fb_font_name in fallback_unicode_fonts:
            for fb_path in colab_font_paths:
                 potential_path = fb_path.replace(font_name, fb_font_name)
                 if os.path.exists(potential_path):
                     try:
                         return ImageFont.truetype(potential_path, default_size)
                     except Exception:
                         pass # Try next fallback
    except Exception:
        pass # Continue to reportlab fallback


    # Fallback to ReportLab's default font if Pillow fails
    print(f"⚠️ Font {font_name} not found or could not be loaded. Using default Pillow font.")
    try:
        # Attempt to return a basic Pillow font if no TTF found
        return ImageFont.load_default()
    except Exception as e:
        print(f"❌ Could not load default Pillow font: {e}")
        raise RuntimeError("Could not load any font.")


# ---------------- Delete after 5 minutes ---------------- #
def schedule_delete(file_path, delay=300):  # 300 sec = 5 minutes
    try:
        def delete_file():
            if os.path.exists(file_path):
                try:
                    if os.path.isdir(file_path):
                        import shutil
                        shutil.rmtree(file_path)
                    else:
                        os.remove(file_path)
                    print(f"🗑 Deleted: {file_path}")
                except Exception as e:
                    print(f"❌ Failed to delete {file_path}: {e}")
        timer = threading.Timer(delay, delete_file)
        timer.start()

    except Exception as e:
      print(f"conversion failed: {e}")


# -------------------------------
# Core Conversion Functions
# -------------------------------

def xls_to_csv(xls_path, csv_path, sheet_name=None):
    try:
        wb = load_workbook(xls_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        with open(csv_path, "w", encoding="utf-8", newline="") as out:
            writer = csv.writer(out)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
        print(f"✅ XLS converted to CSV: {csv_path}")
        return csv_path
    except Exception as e:
        print(f"❌ XLS to CSV failed: {e}")
        return None

def xls_to_doc(xls_path, docx_path, sheet_name=None):
    try:
        wb = load_workbook(xls_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        doc = Document()
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            doc.save(docx_path)
            print(f"⚠️ Empty sheet. Saved empty DOCX: {docx_path}")
            return docx_path

        table = doc.add_table(rows=1, cols=len(rows[0]))
        for i, val in enumerate(rows[0]):
            table.cell(0, i).text = "" if val is None else str(val)

        for r in rows[1:]:
            cells = table.add_row().cells
            for i, val in enumerate(r):
                cells[i].text = "" if val is None else str(val)

        doc.save(docx_path)
        print(f"✅ XLS converted to DOCX: {docx_path}")
        return docx_path
    except Exception as e:
        print(f"❌ XLS to DOCX failed: {e}")
        return None

def xls_to_txt(xls_path, txt_path, delimiter="\t", sheet_name=None):
    try:
        wb = load_workbook(xls_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        with open(txt_path, "w", encoding="utf-8") as out:
            for row in ws.iter_rows(values_only=True):
                out.write(delimiter.join("" if v is None else str(v) for v in row) + "\n")
        print(f"✅ XLS converted to TXT: {txt_path}")
        return txt_path
    except Exception as e:
        print(f"❌ XLS to TXT failed: {e}")
        return None

def xls_to_pdf(xls_path, pdf_path, margin=40, line_gap=14, font="Helvetica", size=10, sheet_name=None):
    try:
        wb = load_workbook(xls_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        c = canvas.Canvas(pdf_path, pagesize=A4)
        c.setFont(font, size)

        y = PAGE_H - margin
        max_chars = 180

        for row in ws.iter_rows(values_only=True):
            line = " | ".join("" if v is None else str(v) for v in row)
            chunks = [line[i:i+max_chars] for i in range(0, len(line), max_chars)] or [" "]
            for chunk in chunks:
                if y < margin:
                    c.showPage()
                    c.setFont(font, size)
                    y = PAGE_H - margin
                c.drawString(margin, y, chunk)
                y -= line_gap

        c.save()
        print(f"✅ XLS converted to PDF: {pdf_path}")
        return pdf_path
    except Exception as e:
        print(f"❌ XLS to PDF failed: {e}")
        return None

def xls_to_json(xls_path, json_path, sheet_name=None):
    try:
        wb = load_workbook(xls_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            with open(json_path, "w", encoding="utf-8") as out:
                out.write("[]")
            print(f"⚠️ Empty sheet. Saved empty JSON: {json_path}")
            return json_path

        headers = [str(h) if h is not None else "" for h in rows[0]]
        data = []
        for r in rows[1:]:
            obj = {headers[i]: ("" if r[i] is None else r[i]) for i in range(len(headers))}
            data.append(obj)

        with open(json_path, "w", encoding="utf-8") as out:
            json.dump(data, out, ensure_ascii=False, indent=2)

        print(f"✅ XLS converted to JSON: {json_path}")
        return json_path
    except Exception as e:
        print(f"❌ XLS to JSON failed: {e}")
        return None

def xls_to_image(xls_path, out_path, sheet_name=None, font_size=12, margin=40, line_height=18,
                 img_width=1200, max_lines_per_img=60, max_safe_height=30000):
    """
    Converts an XLS/XLSX file to PNG image(s).
    Handles large files by splitting into multiple images.
    """
    try:
        if not os.path.exists(xls_path):
            raise FileNotFoundError(f"❌ File not found: {xls_path}")

        wb = load_workbook(xls_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        lines = []
        for row in ws.iter_rows(values_only=True):
            lines.append(" | ".join(str(x) if x is not None else "" for x in row))

        if not lines:
            raise RuntimeError("❌ XLS/XLSX sheet is empty.")

        total_lines = len(lines)
        split_into_multiple = total_lines > max_lines_per_img

        # Ask user if too many lines
        if split_into_multiple:
            print(f"Sheet has {total_lines} lines (>{max_lines_per_img}).")
            print("Choose output option:")
            print("1. Single giant image (may be very tall)")
            print("2. Multiple images (zipped)")
            choice = input("Enter choice (1/2): ").strip()
            if choice == "1":
                split_into_multiple = False
            else:
                split_into_multiple = True

        images = []

        def render_chunk(chunk, filename):
            """Render one chunk of XLS data into an image file."""
            # --- Measure wrapped line count first ---
            total_wrapped_lines = 0
            wrapped_chunks = []
            for line in chunk:
                script = detect_script_simple(line)
                font = get_font_path(script, font_size)
                wrapped = textwrap.wrap(line, width=int((img_width - 2*margin) / (font_size*0.6)))
                if not wrapped:
                    wrapped = [""]
                wrapped_chunks.append((wrapped, font))
                total_wrapped_lines += len(wrapped)

            # --- Now allocate correct height ---
            img_height = margin * 2 + total_wrapped_lines * line_height
            if img_height > max_safe_height:
                print(f"⚠️ Warning: Image height {img_height}px may be too large.")
            img = Image.new("RGB", (img_width, img_height), "white")
            draw = ImageDraw.Draw(img)

            y_offset = margin
            for wrapped, font in wrapped_chunks:
                for w_line in wrapped:
                    draw.text((margin, y_offset), w_line, fill="black", font=font)
                    y_offset += line_height

            img.save(filename, "PNG")
            return filename


        if not split_into_multiple:
            # --- Option 1: single image ---
            if not out_path.lower().endswith(".png"):
                out_path += ".png"
            images.append(render_chunk(lines, out_path))

        else:
            # --- Option 2: multiple images zipped ---
            chunks = [lines[i:i+max_lines_per_img] for i in range(0, total_lines, max_lines_per_img)]
            img_dir = os.path.splitext(out_path)[0] + "_images"
            os.makedirs(img_dir, exist_ok=True)

            for i, chunk in enumerate(chunks):
                img_path = os.path.join(img_dir, f"page_{i+1}.png")
                images.append(render_chunk(chunk, img_path))

            # Zip all pages
            zip_path = os.path.splitext(out_path)[0] + ".zip"
            with zipfile.ZipFile(zip_path, 'w') as zf:
                for img_file in images:
                    zf.write(img_file, os.path.basename(img_file))

            shutil.rmtree(img_dir)  # cleanup
            images = [zip_path]
            out_path = zip_path


        print(f"✅ XLS/XLSX converted to image(s): {out_path}")
        return out_path

    except Exception as e:
        print(f"❌ XLS/XLSX to Image conversion failed: {e}")
        return None


# -------------------------------
# CLI Tool (Modified for Colab)
# -------------------------------
def main():
    # For Colab file handling
    try:
        from google.colab import files
    except ImportError:
        files = None

    while True:
        print("\n==== XLS File Converter =====")
        print("1. Convert XLS/XLSX File")
        print("2. Exit")
        choice = input("Enter choice: ").strip()

        if choice == "1":
            # Upload file (Colab) or enter path (offline)
            if files:
                #uploaded = files.upload()
                #xls_file = list(uploaded.keys())[0]
                xls_file = input("Enter path of XLS/XLSX file: ").strip()
            else:
                xls_file = input("Enter path of XLS/XLSX file: ").strip()

            print("\nSupported conversions: ")
            print("1. XLS → CSV")
            print("2. XLS → TXT")
            print("3. XLS → PDF")
            print("4. XLS → DOCX")
            print("5. XLS → JSON")
            print("6. XLS → IMAGE(PNG)") # Added Image option
            fmt_choice = input("Select target format (1-6): ").strip()

            base, _ = os.path.splitext(xls_file)
            out_file = None
            sheet_name = None # You might want to add an option to select sheet name

            try:
                if fmt_choice == "1":
                    out_file = base + ".csv"
                    xls_to_csv(xls_file, out_file, sheet_name)
                elif fmt_choice == "2":
                    out_file = base + ".txt"
                    xls_to_txt(xls_file, out_file, sheet_name=sheet_name)
                elif fmt_choice == "3":
                    out_file = base + ".pdf"
                    xls_to_pdf(xls_file, out_file, sheet_name=sheet_name)
                elif fmt_choice == "4":
                    out_file = base + ".docx"
                    xls_to_doc(xls_file, out_file, sheet_name)
                elif fmt_choice == "5":
                    out_file = base + ".json"
                    xls_to_json(xls_file, out_file, sheet_name)
                elif fmt_choice == "6":
                    out_file = base + ".png" # Initial suggestion, might change to .zip
                    out_file = xls_to_image(xls_file, out_file, sheet_name=sheet_name)
                else:
                    print("❌ Invalid choice!")
                    continue

                if out_file:
                     print(f"✅ Converted successfully: {out_file}")
                     # Schedule auto-delete after 5 minutes
                     schedule_delete(out_file)
                     # Colab download option
                     if files:
                        if os.path.isdir(out_file): # if output is a directory (images)
                             print(f"💡 Multiple files saved in {out_file}. You may need to zip and download manually.")
                        else:
                             files.download(out_file)


            except Exception as e:
                print("❌ Conversion failed:", e)

        elif choice == "2":
            print("👋 Exiting...")
            break
        else:
            print("❌ Invalid choice!")


if __name__ == "__main__":
    main()