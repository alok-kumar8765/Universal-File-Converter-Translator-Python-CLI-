# csv_converters.py
import csv, json
import pandas as pd
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from docx import Document
import os
import threading
from PIL import Image, ImageDraw, ImageFont
import textwrap
import math
import zipfile
import io
import unicodedata
import shutil # Import shutil for rmtree

# ---------- Page setup ----------
PAGE_W, PAGE_H = A4

# ---------------- Delete after 5 minutes ---------------- #
def schedule_delete(file_path, delay=300):  # 300 sec = 5 minutes
    try:
        def delete_file():
            if os.path.exists(file_path):
                try:
                    if os.path.isdir(file_path):
                        import shutil
                        shutil.rmtree(file_path)
                    else:
                        os.remove(file_path)
                    print(f"🗑 Deleted: {file_path}")
                except Exception as e:
                    print(f"❌ Failed to delete {file_path}: {e}")
        timer = threading.Timer(delay, delete_file)
        timer.start()

    except Exception as e:
      print(f"conversion failed: {e}")

def csv_to_xls(csv_path, xls_path, chunksize=10000):
    try:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Sheet1")
        first = True
        for chunk in pd.read_csv(csv_path, chunksize=chunksize, dtype=str):
            if first:
                ws.append(list(chunk.columns))
                first = False
            for row in chunk.itertuples(index=False, name=None):
                ws.append(list(row))
        wb.save(xls_path)
        return xls_path

    except Exception as e:
        raise RuntimeError(f"❌ XLS conversion failed: {e}")

def csv_to_pdf(csv_path, pdf_path, margin=40, line_gap=14, font="Helvetica", size=10):
    try:
        c = canvas.Canvas(pdf_path, pagesize=A4)
        c.setFont(font, size)
        y = PAGE_H - margin
        max_chars = 180
        with open(csv_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            for row in reader:
                line = " | ".join([str(x) for x in row])
                # simple wrap by characters
                for chunk in [line[i:i+max_chars] for i in range(0, len(line), max_chars)] or [" "]:
                    if y < margin:
                        c.showPage(); c.setFont(font, size); y = PAGE_H - margin
                    c.drawString(margin, y, chunk)
                    y -= line_gap
        c.save()
        return pdf_path

    except Exception as e:
        raise RuntimeError(f"❌ PDF conversion failed: {e}")

def csv_to_doc(csv_path, docx_path):
    try:
        df = pd.read_csv(csv_path, dtype=str)
        doc = Document()
        table = doc.add_table(rows=1, cols=len(df.columns))
        for i, col in enumerate(df.columns): table.cell(0,i).text = col
        for row in df.itertuples(index=False, name=None):
            cells = table.add_row().cells
            for i, val in enumerate(row): cells[i].text = "" if pd.isna(val) else str(val)
        doc.save(docx_path)
        return docx_path

    except Exception as e:
        raise RuntimeError(f"❌ DOCX conversion failed: {e}")

def csv_to_txt(csv_path, txt_path, delimiter="\t", chunksize=10000):
    try:
        first = True
        with open(txt_path, "w", encoding="utf-8") as out:
            for chunk in pd.read_csv(csv_path, dtype=str, chunksize=chunksize):
                if first:
                    out.write(delimiter.join(chunk.columns) + "\n")
                    first = False
                for row in chunk.itertuples(index=False, name=None):
                    out.write(delimiter.join("" if pd.isna(x) else str(x) for x in row) + "\n")
        return txt_path
    except Exception as e:
        raise RuntimeError(f"❌ TXT conversion failed: {e}")

def csv_to_json(csv_path, json_path, chunksize=50000):
    try:
        # Streams rows; writes list-of-objects JSON
        first = True
        with open(json_path, "w", encoding="utf-8") as out:
            out.write("[")
            for chunk in pd.read_csv(csv_path, dtype=str, chunksize=chunksize):
                records = chunk.where(pd.notnull(chunk), None).to_dict(orient="records")
                for rec in records:
                    if not first: out.write(",\n")
                    json.dump(rec, out, ensure_ascii=False)
                    first = False
            out.write("]")
        return json_path

    except Exception as e:
        raise RuntimeError(f"❌ TXT conversion failed: {e}")

# ---------- Script to Font mapping (for image conversion) ----------
FALLBACK_FONTS = {
    "LATIN": "arial.ttf", # Assuming Arial is available or will be handled
    "DEVANAGARI": "NotoSansDevanagari-Medium.ttf", # Assuming Noto is available or will be handled
    "CJK": "arialuni.ttf", # Arial Unicode MS
    "ARABIC": "arial.ttf",
    "GREEK": "arial.ttf",
    "OTHER": "arialuni.ttf",
    "DEFAULT": "arial.ttf"
}

def detect_script_simple(text: str) -> str:
    """
    Heuristic script detection based on Unicode codepoints.
    Returns one of: DEVANAGARI, CJK, ARABIC, GREEK, HANGUL, LATIN, OTHER
    """
    if not text:
        return "LATIN"
    for ch in text:
        cp = ord(ch)
        if 0x0900 <= cp <= 0x097F:  # Devanagari (Hindi, etc)
            return "DEVANAGARI"
        if (0x4E00 <= cp <= 0x9FFF) or (0x3400 <= cp <= 0x4DBF) or (0x3040 <= cp <= 0x30FF):
            return "CJK"
        if 0xAC00 <= cp <= 0xD7AF:
            return "HANGUL"
        if 0x0600 <= cp <= 0x06FF or 0x0750 <= cp <= 0x077F:
            return "ARABIC"
        if 0x0370 <= cp <= 0x03FF:
            return "GREEK"
    if any(ord(ch) > 127 for ch in text):
        return "OTHER"
    return "LATIN"

def get_font_path(script, default_size=12):
    font_name = FALLBACK_FONTS.get(script, FALLBACK_FONTS["DEFAULT"])
    # In a Colab environment, commonly available fonts are in specific paths
    # You might need to adjust these paths or download fonts if needed.
    # Example paths (may vary):
    colab_font_paths = [
        f"/usr/share/fonts/truetype/liberation/{font_name}", # Liberation fonts often present
        f"/usr/share/fonts/truetype/msttcorefonts/{font_name}", # If ttf-mscorefonts-installer is run
        f"/usr/share/fonts/truetype/noto/{font_name}", # Noto fonts
        f"/content/{font_name}", # If user uploaded
        font_name # As a last resort, try just the name
    ]

    for font_path in colab_font_paths:
        if os.path.exists(font_path):
            try:
                return ImageFont.truetype(font_path, default_size)
            except Exception as e:
                print(f"⚠️ Could not load font {font_path}: {e}")

    # If specific font not found, try a generic fallback font that supports Unicode
    try:
        # Try a widely available Unicode font like Arial Unicode MS or Noto Sans
        fallback_unicode_fonts = ["arialuni.ttf", "NotoSans-Regular.ttf"]
        for fb_font_name in fallback_unicode_fonts:
            for fb_path in colab_font_paths:
                 potential_path = fb_path.replace(font_name, fb_font_name)
                 if os.path.exists(potential_path):
                     try:
                         return ImageFont.truetype(potential_path, default_size)
                     except Exception:
                         pass # Try next fallback
    except Exception:
        pass # Continue to reportlab fallback

    # Fallback to ReportLab's default font if Pillow fails
    print(f"⚠️ Font {font_name} not found or could not be loaded. Using default Pillow font.")
    try:
        # Attempt to return a basic Pillow font if no TTF found
        return ImageFont.load_default()
    except Exception as e:
        print(f"❌ Could not load default Pillow font: {e}")
        raise RuntimeError("Could not load any font.")


def csv_to_image(csv_path, out_path, font_size=12, margin=40, line_height=18,
                 img_width=1200, max_lines_per_img=60, max_safe_height=30000):
    """
    Converts a CSV file to PNG image(s).
    Handles large files by splitting into multiple images.
    Fix: image height now accounts for wrapped lines (no cropping).
    """
    try:
        if not os.path.exists(csv_path):
            raise FileNotFoundError(f"❌ File not found: {csv_path}")

        # Read CSV lines
        lines = []
        with open(csv_path, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for row in reader:
                lines.append(" | ".join(str(x) for x in row))

        if not lines:
            raise RuntimeError("❌ CSV file is empty.")

        total_lines = len(lines)
        split_into_multiple = total_lines > max_lines_per_img

        # Ask user if too many lines
        if split_into_multiple:
            print(f"CSV has {total_lines} lines (>{max_lines_per_img}).")
            print("Choose output option:")
            print("1. Single giant image (may be very tall)")
            print("2. Multiple images (zipped)")
            choice = input("Enter choice (1/2): ").strip()
            if choice == "1":
                split_into_multiple = False
            else:
                split_into_multiple = True

        images = []

        def render_chunk(chunk, filename):
            """Render one chunk of CSV into an image file."""
            # --- Measure wrapped line count first ---
            total_wrapped_lines = 0
            wrapped_chunks = []
            for line in chunk:
                script = detect_script_simple(line)
                font = get_font_path(script, font_size)
                wrapped = textwrap.wrap(line, width=int((img_width - 2*margin) / (font_size*0.6)))
                if not wrapped:
                    wrapped = [""]
                wrapped_chunks.append((wrapped, font))
                total_wrapped_lines += len(wrapped)

            # --- Now allocate correct height ---
            img_height = margin * 2 + total_wrapped_lines * line_height
            if img_height > max_safe_height:
                print(f"⚠️ Warning: Image height {img_height}px may be too large.")
            img = Image.new("RGB", (img_width, img_height), "white")
            draw = ImageDraw.Draw(img)

            y_offset = margin
            for wrapped, font in wrapped_chunks:
                for w_line in wrapped:
                    draw.text((margin, y_offset), w_line, fill="black", font=font)
                    y_offset += line_height

            img.save(filename, "PNG")
            return filename

        if not split_into_multiple:
            # --- Option 1: single image ---
            if not out_path.lower().endswith(".png"):
                out_path += ".png"
            images.append(render_chunk(lines, out_path))

        else:
            # --- Option 2: multiple images zipped ---
            chunks = [lines[i:i+max_lines_per_img] for i in range(0, total_lines, max_lines_per_img)]
            img_dir = os.path.splitext(out_path)[0] + "_images"
            os.makedirs(img_dir, exist_ok=True)

            for i, chunk in enumerate(chunks):
                img_path = os.path.join(img_dir, f"page_{i+1}.png")
                images.append(render_chunk(chunk, img_path))

            # Zip all pages
            zip_path = os.path.splitext(out_path)[0] + ".zip"
            with zipfile.ZipFile(zip_path, 'w') as zf:
                for img_file in images:
                    zf.write(img_file, os.path.basename(img_file))

            shutil.rmtree(img_dir)  # cleanup
            images = [zip_path]
            out_path = zip_path

        print(f"✅ CSV converted to image(s): {out_path}")
        return out_path

    except Exception as e:
        print(f"❌ CSV to Image conversion failed: {e}")
        return None

# ---------------- CLI MENU ---------------- #

def main():
    # For Colab file handling
    try:
        from google.colab import files
    except ImportError:
        files = None

    while True:
        print("\n==== CSV File Converter =====")
        print("1. Convert CSV File")
        print("2. Exit")
        choice = input("Enter choice: ").strip()

        if choice == "1":
            # Upload file (Colab) or enter path (offline)
            if files:
                uploaded = files.upload()
                csv_file = list(uploaded.keys())[0]
                #csv_file = input("Enter path of CSV file: ").strip()
            else:
                csv_file = input("Enter path of CSV file: ").strip()


            print("\nSupported conversions: ")
            print("1. CSV → PDF")
            print("2. CSV → TXT")
            print("3. CSV → PNG (image)")
            print("4. CSV → DOCX")
            print("5. CSV → XLSX")
            print("6. CSV → JSON")
            fmt_choice = input("Select target format (1-6): ").strip()

            base, _ = os.path.splitext(csv_file)
            out_file = None

            try:
                if fmt_choice == "1":
                    out_file = base + ".pdf"
                    csv_to_pdf(csv_file, out_file)
                elif fmt_choice == "2":
                    out_file = base + ".txt"
                    csv_to_txt(csv_file, out_file)
                elif fmt_choice == "3":
                    out_file = base + ".png" # Initial suggestion, might change to .zip
                    out_file = csv_to_image(csv_file, out_file) # Update out_file with the actual output path
                elif fmt_choice == "4":
                    out_file = base + ".docx"
                    csv_to_doc(csv_file, out_file)
                elif fmt_choice == "5":
                    out_file = base + ".xlsx"
                    csv_to_xls(csv_file, out_file)
                elif fmt_choice == "6":
                    out_file = base + ".json"
                    csv_to_json(csv_file, out_file)
                else:
                    print("❌ Invalid choice!")
                    continue

                if out_file:
                     print(f"✅ Converted successfully: {out_file}")
                     # Schedule auto-delete after 5 minutes
                     schedule_delete(out_file)
                     # Colab download option
                     if files:
                        if os.path.isdir(out_file): # if output is a directory (images)
                             print(f"💡 Multiple files saved in {out_file}. You may need to zip and download manually.")
                             # Example of zipping and offering download (requires zip installed)
                             # !zip -r {out_file}.zip {out_file}
                             # files.download(f"{out_file}.zip")
                        else:
                             files.download(out_file)


            except Exception as e:
                print("❌ Conversion failed:", e)

        elif choice == "2":
            print("👋 Exiting...")
            break
        else:
            print("❌ Invalid choice!")

if __name__ == "__main__":
    main()